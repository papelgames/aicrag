from app.models import Productos, Proveedores, MensajesSistema, TareasSistema, CabecerasPresupuestos, Estados
from flask import current_app
from app import create_app
import os
from werkzeug.utils import secure_filename
from time import strftime, gmtime
import locale
from dotenv import load_dotenv
import openpyxl
from rq import get_current_job
from datetime import datetime

load_dotenv()
settings_module = os.getenv('APP_SETTINGS_MODULE')
app = create_app(settings_module)

app.app_context().push()

def sin_codigo_barras_to_excel():
   #actualizo el inicio de la tarea
   job = get_current_job()
   tarea = TareasSistema.get_by_id_rq(job.id)
   tarea.fecha_inicio = datetime.now()
   tarea.save()
   
   # Filtrar y preparar los datos para la inserción
   archivo_dir = current_app.config['ARCHIVOS_PARA_DESCARGA']
   productos_incompletos = Productos.get_all_productos_sin_codigo_de_barras()

   wb = openpyxl.Workbook()
   hoja = wb.active
   hoja.append(('Codigo_de_barras', 'Id_proveedores', 'Descripcion', 'Importe', 'Nombre_proveedor'))
   total = len(productos_incompletos)
   registros_error = 0
   for i, row in enumerate(productos_incompletos):
      try:
         if job:
            job.meta['progress'] = 100 * i / total
            job.save_meta()
         hoja.append((row[0],row[1],row[2],row[3],row[4]))
      except Exception as e:
         registros_error += 1
         print (f"error de registro {i} id {id[0].value}: {e} ") #llevarlo a logger
   wb.save(archivo_dir + '/productos_sin_codigo_barra.xlsx')
   mensaje = MensajesSistema(asunto="Exportación prod. sin cod. barras", cuerpo=f"El archivo con los productos sin códigos de barra se exportó correctamente. Total de registros: {total}. Registros con error: {registros_error}")
   mensaje.save()
   tarea.complete = True
   tarea.save()
   if job:
      job.meta['progress'] = 100
      job.save_meta()


def to_precios_dbf():
     #actualizo el inicio de la tarea
   job = get_current_job()
   tarea = TareasSistema.get_by_id_rq(job.id)
   tarea.fecha_inicio = datetime.now()
   tarea.save()
  # Filtrar y preparar los datos para la inserción
   archivo_dir = current_app.config['ARCHIVOS_PARA_DESCARGA']
   productos_precios = Productos.get_all_precios_dbf()
   table = [(f'CODIGO,C,13\tDETALLE,C,56\tPRECIOVP,N,10,4')]
   errores = []
   locale.setlocale(locale.LC_ALL, '')
   total=len(productos_precios)
   registros_error = 0
   print (f"Total es: {total}")
   #recorro la tabla y si ha precios que no concuerdan con el tipo de precio que soporta el dbf armo un archivo txt
   #para que el usuario pueda revisar esos productos y corregirlos. 

   for i, row in enumerate(productos_precios):
      try:
         if job:
            job.meta['progress'] = 100 * i / total
            job.save_meta()
                           
         if len(str(round(row[2]))) > 10:
            errores.append(f'{row[0][:13]}\t{row[1][:56]}\t{row[2]}')
         else:
            table.append(f'{row[0][:13]}\t{row[1][:56]}\t{locale._format("%.2f", row[2], grouping=True, monetary=True)}')
      except Exception as e:
         registros_error +=1
         print (f"error de registro {i} id {id[0].value}: {e} ") #llevarlo a logger

   # Inserto los importes no soportados en el archivo precios_elevados.txt siempre y cuando haya.
   if errores:
      with open(archivo_dir + '/precios_elevados.txt', 'w') as errores_file:
         errores_file.write('\n'.join(errores))
   
   # Inserto los datos en el archivo .dbf
   with open(archivo_dir + '/Precios.dbf', 'w') as precios_file:
      precios_file.write('\n'.join(table))
   
   mensaje = MensajesSistema(asunto="Exportación archivo .dbf", cuerpo=f"El arhivo .dbf se generó correctamente con {total} y tuvo {registros_error} registros con error")
   mensaje.save()
   tarea.complete = True
   tarea.save()
   if job:
      job.meta['progress'] = 100
      job.save_meta()

def in_lista_masiva(file_path, id_proveedor, user):
   proveedor = Proveedores.get_by_id(id_proveedor)

   #actualizo el inicio de la tarea
   job = get_current_job()
   tarea = TareasSistema.get_by_id_rq(job.id)
   tarea.fecha_inicio = datetime.now()
   tarea.save()
   
   #validación migrada
   #abro documento excel
   import openpyxl 
   documento = openpyxl.load_workbook(os.path.abspath(file_path), data_only= True)
   ws = documento.active
   
   #traigo parametria del proveedor
   columnas = [proveedor.nombre,
               proveedor.formato_id,
               proveedor.columna_id_lista_proveedor, 
               proveedor.columna_codigo_de_barras, 
               proveedor.columna_descripcion,
               proveedor.columna_importe,
               proveedor.columna_utilidad ]
   
   #indico que al excel en que columnas el proveedor carga cada dato.
   rango_id_lista_proveedor =  ws[columnas[2]]
   nombre_archivo = os.path.basename(file_path)
   secuencia = 0
   control_proveedor = False
   # controlo que el archivo corresponda al proveedor
   for id in rango_id_lista_proveedor:
         if secuencia == 15:
               break
         if str(id.value).upper() == str(columnas[1]).upper():
               control_proveedor = True
               break
         secuencia +=1
   if control_proveedor == True:
      #indico que al excel en que columnas el proveedor carga cada dato.
      rango_id_lista_proveedor =  ws[columnas[2]]
      rango_codigo_de_barras =  ws[columnas[3]]
      rango_descripcion =  ws[columnas[4]]
      rango_importe =  ws[columnas[5]]
      rango_utilidad = ws[columnas[6]]
      #incremental de cada caso
      registros_nuevos = 0
      registros_actualizados = 0
      registros_total = 0
      registros_ignorados = 0
      registros_error = 0
      #genero un id unico por subida para cada registro
      id_ingreso = str(strftime('%d%m%y%H%m%s', gmtime()))
      #creo una matriz con los datos del excel para luego iterarla.   
      mat = list(zip(rango_id_lista_proveedor, rango_codigo_de_barras, rango_descripcion, rango_importe, rango_utilidad))
      #calculo total para poder calcular porcentaje de la tarea.
      total = len(mat)
      #inserto los registros que no existen
      producto_nuevo = Productos()
      producto_por_id=None
      
      for i, id in enumerate(mat):
         try:   
            if id[0].value != None and str(id[0].value).upper() != str(columnas[1]).upper():
               producto_por_id = Productos.get_by_id_lista_proveedor(id[0].value, id_proveedor ) #corregir esta consulta
               registros_total += 1
               
               if job:
                  job.meta['progress'] = 100 * i / total
                  job.save_meta()
                           
               #si es un producto nuevo
               if not producto_por_id:
                  if id[4].value == None:
                        utilidad_ = 100
                  else:
                        utilidad_ = id[4].value
                  #antes de grabar chequeo si el proveedor guarda con iva o no
                  if proveedor.incluye_iva == True:
                        producto_nuevo = Productos(codigo_de_barras = id[1].value,
                                                id_proveedor = id_proveedor,
                                                id_lista_proveedor = id[0].value,
                                                descripcion = id[2].value,
                                                importe = round(id[3].value,2),
                                                utilidad = utilidad_,
                                                cantidad_presentacion = 1,
                                                id_ingreso = id_ingreso,
                                                es_servicio = False,
                                                usuario_alta = user,
                                                usuario_modificacion = user
                                                )
                        registros_nuevos += 1 
                  else:
                        producto_nuevo = Productos(codigo_de_barras = id[1].value,
                                                   id_proveedor = id_proveedor,
                                                   id_lista_proveedor = id[0].value,
                                                   descripcion = id[2].value,
                                                   importe = round(id[3].value * 1.21 ,2),
                                                   utilidad = utilidad_,
                                                   cantidad_presentacion = 1,
                                                   id_ingreso = id_ingreso,
                                                   es_servicio = False,
                                                   usuario_alta = user,
                                                   usuario_modificacion = user
                                                   )
                        registros_nuevos += 1
                  producto_nuevo.only_add()
               
               #actualizo productos que existe si es que tienen un importe distinto al cargado.    
               if producto_por_id:
                  #chequeo si el codigo de barras cambió o si lo agregron
                  if producto_por_id.codigo_de_barras == None and id[1].value != None:
                     producto_por_id.codigo_de_barras = id[1].value
                     producto_por_id.id_ingreso = id_ingreso
                     producto_por_id.only_add()
                  elif producto_por_id.codigo_de_barras != str(id[1].value) and id[1].value != None:
                     producto_por_id.codigo_de_barras = id[1].value
                     producto_por_id.id_ingreso = id_ingreso
                     producto_por_id.only_add()
                  if proveedor.incluye_iva == True:
                     if float(producto_por_id.importe) != round(id[3].value,2):    
                        producto_por_id.importe = round(id[3].value,2)
                        producto_por_id.usuario_modificacion = user
                        producto_por_id.id_ingreso = id_ingreso
                        registros_actualizados += 1   
                        producto_por_id.only_add()
                     else:
                        registros_ignorados += 1 
                  #si el proveedor pasa la lista sin iva
                  else:
                     if float(producto_por_id.importe) != round(id[3].value * 1.21 ,2):
                        producto_por_id.importe = round(id[3].value * 1.21 ,2)
                        producto_por_id.usuario_modificacion = user
                        producto_por_id.id_ingreso = id_ingreso
                        registros_actualizados += 1
                        producto_por_id.only_add()
                     else:
                        registros_ignorados += 1
         except Exception as e:
            registros_error += 1
            print (f"error de registro {i} id {id[0].value}: {e} ")
                  
      #commiteo las tablas
      if producto_por_id:
         producto_por_id.save()
      producto_nuevo.only_save()
      mensaje = MensajesSistema(asunto="Importación masiva ok", 
                                cuerpo=f"La importación de {proveedor.nombre} masiva terminó correctamente con el archivo {nombre_archivo} con registros nuevos: {registros_nuevos}, registros actualizados: {registros_actualizados}, registros ignorados: {registros_ignorados} y registros con error:{registros_error} ")
      mensaje.save()
      tarea.complete = True
      tarea.save()
      if job:
        job.meta['progress'] = 100
        job.save_meta()
   elif control_proveedor == False:
      mensaje = MensajesSistema(asunto="Importación masiva con error", 
                                cuerpo=f"El archivo {nombre_archivo} no corresponde al proveedor {proveedor.nombre}. La importación no se realizó", 
                                alerta=True)
      mensaje.save()
      tarea.error = True
      tarea.complete = True
      tarea.save()

def actualiza_estado_presupuestos():
   estado_pendiente = Estados.get_first_by_clave_tabla(1,"estado_presupuesto")
   presupuestos = CabecerasPresupuestos.get_all_estado(estado_pendiente.id)
   job = get_current_job()
   tarea = TareasSistema.get_by_id_rq(job.id)
   tarea.fecha_inicio = datetime.now()
   tarea.save()
   total = 0
   registros_error = 0
   estado_vencido = Estados.get_first_by_clave_tabla(2,"estado_presupuesto")
   for i, vencimiento in enumerate(presupuestos):
      total += 1
      try:
         if vencimiento.fecha_vencimiento < datetime.now():
            print (f'Id del presupuesto vencido {vencimiento.id}')
            vencimiento.id_estado = estado_vencido.id
            vencimiento.save()

            if job:
               job.meta['progress'] = 100 * i / total
               job.save_meta()
      except Exception as e:
            registros_error += 1
            print (f"Error de registro {i} id {id[0].value}: {e} ")
   
   mensaje = MensajesSistema(asunto="Actualización estados presupuestos", 
                             cuerpo=f"Se han actualizado {total} registros a vencido", 
                             leido=True)
   mensaje.save()
   tarea.complete = True
   tarea.save()
   if job:
      job.meta['progress'] = 100
      job.save_meta()